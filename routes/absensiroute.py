from openpyxl import load_workbook ,Workbook
from datetime import datetime

def insertdb(scanned_name):
    file_path = 'presence.xlsx'
    
    try:
        # Try to load the existing workbook
        wb = load_workbook(file_path)
        sheet = wb.active
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        sheet = wb.active
        sheet['A1'] = 'Nama'
        sheet['B1'] = 'Tanggal dan Waktu Scan'

    next_row = sheet.max_row + 1
    date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Insert data
    sheet[f'A{next_row}'] = scanned_name
    sheet[f'B{next_row}'] = date

    # Save the workbook
    wb.save(file_path)

    print("Data successfully entered and saved to 'data_entry2.xlsx'")